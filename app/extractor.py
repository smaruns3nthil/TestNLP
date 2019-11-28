# Import pandas
import pandas as pd
import spacy

from openpyxl import load_workbook

#list of stop triggers
negationtriggers=["no", "not","denies","nothing","asymptomatic","atypical","deny","denied","denying","free","lack","negative","never","without","nor"]

hypothetical=["believes","believe","agree","agree","agreement","favor","favoured","outline","outlined","likely","risk","possible","typical","typically","candidate",
              "regarding","discuss","discussion","publication","advised","advise","advising","agreeable","agreed","allows","discussed","amenable","anticipate",
              "approximately","recommend","recommended","recommending","ask","authorization","await","initiation","chance","concern","concerened","concerning","consent"
              "could","can","criteria","desire","desired","desires","educate","meant","prefer","prefers","emphasize","emphasized","evaluate","express","expressed","extrapolate",
              "fear","felt","feel","example","information","conversation","decide","decided","declined","elected","wondering","thought","stated","opinion","likelihood",
              "probability","suspicious","hoping","beneficial","usually","reasonable","willing","leaning","likehood","consider","considered","feasible"
              "indicated","recommended","recommends","interested","trying","contemplating","suggests","suspect","suspects","potentially","potential","hesitate"]

family=["sister","daughter","father","brother","cousin","aunt","uncle","family"]


# Load in the workbook
wb = load_workbook('sentences.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')

for i in range(1, 23564):
     #print(str(sheet.cell(row=i, column=1).value))
     #print(i, sheet.cell(row=i, column=1).value)

     nlp= spacy.load('en_core_web_sm')
     doc=nlp(unicode(sheet.cell(row=i, column=1).value),"utf-8")
     parent=''
     type=''
     text=''
     lemma=''
     pos=''
     tag=''
     dep=''
     shape=''
     stopis=False
     for token in doc:
        if token.text in negationtriggers:
            #print(token.text)
            parent=token.head.text
            #print(parent)
            type="negation"
            text=token.text
            lemma=token.lemma_
            pos=token.pos_
            tag=token.tag_
            dep=token.dep_
            shape=token.shape_
            stopis=token.is_stop
        elif token.text in hypothetical:
            #print(token.text)
            parent=token.head.text
            #print(parent)
            type="hypothetical"
            text=token.text
            lemma=token.lemma_
            pos=token.pos_
            tag=token.tag_
            dep=token.dep_
            shape=token.shape_
            stopis=token.is_stop

        elif token.text in family:
            #print(token.text)
            parent=token.head.text
            #print(parent)
            type="family"
            text=token.text
            lemma=token.lemma_
            pos=token.pos_
            tag=token.tag_
            dep=token.dep_
            shape=token.shape_
            stopis=token.is_stop

     sheet.cell(row=i,column=2).value=parent
     sheet.cell(row=i,column=3).value=type
     sheet.cell(row=i,column=4).value=text
     sheet.cell(row=i,column=5).value=lemma
     sheet.cell(row=i,column=6).value=pos
     sheet.cell(row=i,column=7).value=tag
     sheet.cell(row=i,column=8).value=dep
     sheet.cell(row=i,column=9).value=shape
     sheet.cell(row=i,column=10).value=stopis

df = pd.DataFrame(sheet.values)
    # Specify a writer
writer = pd.ExcelWriter('spacyparsedall.xlsx', engine='xlsxwriter')

# Write your DataFrame to a file     
df.to_excel(writer, 'Sheet1')

# Save the result 
writer.save()
